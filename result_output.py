import matplotlib.pyplot as plt
import matplotlib.colors
import numpy as np
from openpyxl import Workbook
from numpy import mean
from random import randint


def draw_figure(dict_):
    plt.style.use('ggplot')

    for figure_name in dict_.keys():
        dict_[figure_name]['color'] = np.random.rand(3)

        plt.plot([i for i in dict_[figure_name].keys()],
                 [np.mean(i) for i in dict_[figure_name].values()],
                 'o-', color=dict_[figure_name]['color'])

        plt.xticks(np.asarray([i for i in dict_[figure_name].keys()]))
        plt.xlabel('# of elements (1000\'s)')
        plt.ylabel('Time (seconds)')
        plt.title('{0} Figure'.format(figure_name))

        plt.savefig(figure_name)
        plt.close()

    # plt.style.use('ggplot')

    for figure_name in dict_.keys():

        plt.plot([i for i in dict_[figure_name].keys()[:-1]],
                 [np.mean(i) for i in dict_[figure_name].values()[:-1]],
                 'o-', color=dict_[figure_name]['color'],
                 label=figure_name)
        plt.xticks(np.asarray([i for i in dict_[figure_name].keys()[:-1]]))
        dict_[figure_name].pop('color')

    plt.xlabel('# of elements (1000\'s)')
    plt.ylabel('Time (seconds)')
    plt.title('Sorting Analytics Figure')
    plt.legend()
    plt.savefig('Sorting Analytics Figure')
    plt.close()


def write_xlxs(dict_):
    wb = Workbook()

    for sheet_name in dict_.keys():
        sheet = wb.create_sheet(sheet_name)

        c_count = 1
        for N in dict_[sheet_name].keys():
            sheet.cell(row=1, column=c_count, value=N)
            r_count = 2
            for vals in dict_[sheet_name][N]:
                sheet.cell(row=r_count, column=c_count, value=vals)
                r_count += 1
            c_count += 1

    wb.remove(wb['Sheet'])
    wb.save(filename='Sorting_analytics.xlsx')
